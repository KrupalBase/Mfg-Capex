"""
Shared utilities for the CAPEX pipeline.

- PO data cleaning and formatting for CSV export
- Ramp CC transaction normalization
- BF1 station loading from Excel
- Product category splitting and section header merging
- 3-tier auto-mapping agent for station assignment
- Part number extraction from descriptions
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import pandas as pd

# ---------------------------------------------------------------------------
# Text normalization helpers (unchanged from original)
# ---------------------------------------------------------------------------

_NEWLINE_PATTERN = re.compile(r"[\r\n\u2028\u2029]+")


def _to_single_line(text: Any) -> str:
    """Replace any newline variant with space and collapse whitespace."""
    if pd.isna(text):
        return ""
    s = str(text).strip()
    s = _NEWLINE_PATTERN.sub(" ", s)
    return re.sub(r"\s+", " ", s).strip()


def _strip_html(text: Any) -> str:
    if pd.isna(text) or not isinstance(text, str):
        return "" if pd.isna(text) else str(text)
    s = re.sub(r"<[^>]+>", " ", text).replace("&nbsp;", " ").replace("&amp;", "&")
    return re.sub(r"\s+", " ", s).strip()


def _extract_en_us_name(value: Any) -> str:
    """Extract en_US display name from Odoo JSON name field."""
    if pd.isna(value):
        return ""
    if not isinstance(value, str):
        return str(value)
    s = value.strip()
    if not s or (s[0] != "{" and "en_US" not in s):
        return s
    try:
        data = json.loads(value.replace("'", '"'))
        if isinstance(data, dict) and "en_US" in data:
            return str(data["en_US"]).strip()
        return value
    except (json.JSONDecodeError, TypeError):
        return value


def _format_currency(val: Any) -> Any:
    if pd.isna(val):
        return ""
    try:
        n = float(val)
        return round(n, 2) if n == n else val
    except (TypeError, ValueError):
        return val


def _format_qty(val: Any) -> Any:
    if pd.isna(val):
        return ""
    try:
        n = float(val)
        if n == int(n):
            return int(n)
        return round(n, 4)
    except (TypeError, ValueError):
        return val


def _format_ts(val: Any) -> str:
    """Format timestamp to YYYY-MM-DD HH:MM for readability."""
    if pd.isna(val):
        return ""
    try:
        ts = pd.Timestamp(val)
        return ts.strftime("%Y-%m-%d %H:%M")
    except (TypeError, ValueError):
        return str(val)


def clean_po_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Clean and format PO line-level dataframe for CSV export."""
    if df.empty:
        return df
    out = df.copy()

    if "project_name" in out.columns:
        out["project_name"] = out["project_name"].apply(_extract_en_us_name)
    if "po_notes" in out.columns:
        out["po_notes"] = out["po_notes"].apply(_strip_html)
    if "line_description" in out.columns:
        out["line_description"] = out["line_description"].apply(_to_single_line)
    if "po_notes" in out.columns:
        out["po_notes"] = out["po_notes"].apply(_to_single_line)

    money_cols = [
        "price_unit", "price_subtotal", "price_tax", "price_total",
        "po_amount_untaxed", "po_amount_tax", "po_amount_total",
        "bill_amount_total", "bill_amount_paid", "bill_amount_open",
    ]
    for c in money_cols:
        if c in out.columns:
            out[c] = out[c].apply(_format_currency)
    for c in ("product_qty", "qty_received"):
        if c in out.columns:
            out[c] = out[c].apply(_format_qty)

    date_cols = ["date_order", "date_approve", "line_date_planned",
                 "po_created_date", "po_updated_date"]
    for c in date_cols:
        if c in out.columns:
            out[c] = out[c].apply(_format_ts)

    return out


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

KNOWN_CATEGORIES: list[str] = sorted([
    "Non-Inventory: Construction in Process",
    "Non-Inventory: Machinery >$2k",
    "Non-Inventory: Machinery <$2k",
    "Non-Inventory: Furniture >$2k",
    "Non-Inventory: Furniture Fixture / Workstation",
    "Non-Inventory: R&D Parts",
    "Non-Inventory: R&D Services",
    "Non-Inventory: R&D Shipping & Postage",
    "Non-Inventory: R&D Testing Equipment >$2k",
    "Non-Inventory: Software and Applications",
    "Non-Inventory: Tooling and Consumables",
    "Non-Inventory: Shop Tooling and Consumables",
    "Non-Inventory: Deployment Tooling & Supplies",
    "Non-Inventory: IT Equipment >$2k",
    "Non-Inventory: Office Equipment & Supplies",
    "Non-Inventory: Inbound Production Shipping",
    "Non-Inventory: G&A Shipping & Logistics",
], key=len, reverse=True)

CAPEX_CATEGORIES: set[str] = {
    "Non-Inventory: Construction in Process",
    "Non-Inventory: Machinery >$2k",
    "Non-Inventory: Machinery <$2k",
    "Non-Inventory: Furniture >$2k",
    "Non-Inventory: Furniture Fixture / Workstation",
    "Non-Inventory: R&D Testing Equipment >$2k",
    "Non-Inventory: IT Equipment >$2k",
    "Non-Inventory: Software and Applications",
}

RAMP_TO_ODOO_CATEGORY: dict[str, str] = {
    "Research & Development:R&D Materials": "Non-Inventory: R&D Parts",
    "Facility Expense:Tooling & Consumables": "Non-Inventory: Tooling and Consumables",
    "R&D Equipment >$2k": "Non-Inventory: R&D Testing Equipment >$2k",
    "Machinery >$2k": "Non-Inventory: Machinery >$2k",
    "Furniture & Fixtures >$2K": "Non-Inventory: Furniture >$2k",
    "Facility Expense:Furniture & Fixtures <$2K": "Non-Inventory: Furniture Fixture / Workstation",
    "Construction in Process - Internal": "Non-Inventory: Construction in Process",
    "IT Equipment >$2k": "Non-Inventory: IT Equipment >$2k",
    "Facility Expense:Office Equipment & Supplies": "Non-Inventory: Office Equipment & Supplies",
    "Other General & Administrative Costs:Software & Apps": "Non-Inventory: Software and Applications",
    "Facility Expense:Repair & Maintenance": "Non-Inventory: Shop Tooling and Consumables",
    "Other General & Administrative Costs:IT Equipment <$2k": "Non-Inventory: Machinery <$2k",
    "Other General & Administrative Costs:G&A Shipping & Postage": "Non-Inventory: G&A Shipping & Logistics",
}

RAMP_USER_ALIASES: dict[str, str] = {
    "Andrew Ross": "Andy Ross",
    "Chris George": "Christopher George",
    "Chris Johnston": "Chris Johnston",
    "Eduardo Martinez V": "Eduardo Martinez V.",
}

NON_PROD_PROJECTS: set[str] = {
    "BF1-NPI & Pilot Equipment",
    "BF1-Quality Equipment",
    "BF1-Facilities and Infrastructure",
    "BF1-Manufacturing IT Systems",
    "BF1-Warehousing and Material Handling",
    "BF1-Maintenance and Spares",
    "BF1-Prototype R&D Lines",
    "BF1-Other Allocation",
}

PILOT_RAMP_CARDS: set[str] = {
    "B2 test equipment",
    "Builds materials for lab work",
    "General purchasing of R&D manufacturing equipment",
}

_PAYMENT_TERMS_PATTERN = re.compile(
    r"(payment\s+terms?|down\s+payment|deposit|100%\s+due|"
    r"purchase\s+order\s+is\s+governed|T&Cs?\s+dated)",
    re.IGNORECASE,
)


# ---------------------------------------------------------------------------
# Item bucket classification for the spares catalog
# ---------------------------------------------------------------------------

_BUCKET_RULES: list[tuple[str, re.Pattern[str]]] = [
    ("Discount / Credit", re.compile(
        r"(discount|rebate|credit\s*back)", re.IGNORECASE)),
    ("Tariff / Surcharge", re.compile(
        r"(tariff|surcharge|(?<!heavy\s)(?<!medium\s)(?<!light\s)duty\b|"
        r"duties\b|customs\b)", re.IGNORECASE)),
    ("Shipping / Logistics", re.compile(
        r"(^shipping|^freight|^crating|^handling|^DDP\b|ship\s*&\s*handl|"
        r"shipping\s*(cost|charge|&)|unit\s+shipping|inbound.*shipping|"
        r"postage|logistics)", re.IGNORECASE)),
    ("Warranty / Support", re.compile(
        r"(warranty|support\s+package|tech\s+support|maintenance\s+contract|"
        r"service\s+agreement|service\s+contract)", re.IGNORECASE)),
    ("Training", re.compile(
        r"(training\b)", re.IGNORECASE)),
    ("Rental / Lease", re.compile(
        r"(rental|lease\b|monthly\s+rental)", re.IGNORECASE)),
    ("Permitting / Compliance", re.compile(
        r"(^permitting|^permit\b|compliance\b|certification\b)", re.IGNORECASE)),
    ("Services / Labor", re.compile(
        r"(^labor\b|^line\s+\d+\s+labor|^installation\b|^install\b|"
        r"on-?site\s+install|commissioning|startup|start-?up|"
        r"^project\s+management|^engineering\s+hour|concepting\s+hours|"
        r"process\s+engineering|^consulting|^FAT\b.*\bSAT\b|"
        r"integration\s+package|controls\s+integration|"
        r"electrical\s+installation|mechanical.*installation|"
        r"on-?site\s+fee|calibration\s+service)", re.IGNORECASE)),
    ("Software", re.compile(
        r"(^software\b|fleet\s+management\s+software|license\s+key|"
        r"subscription\b)", re.IGNORECASE)),
]

_CATEGORY_BUCKET_MAP: dict[str, str] = {
    "Non-Inventory: R&D Services": "Services / Labor",
    "Non-Inventory: R&D Shipping & Postage": "Shipping / Logistics",
    "Non-Inventory: G&A Shipping & Logistics": "Shipping / Logistics",
    "Non-Inventory: Inbound Production Shipping": "Shipping / Logistics",
    "Non-Inventory: Software and Applications": "Software",
}

CAPITAL_EQUIPMENT_THRESHOLD = 50_000.0


def classify_item_bucket(
    description: str,
    product_category: str,
    avg_unit_price: float,
    total_spend: float,
) -> str:
    """Classify a spares-catalog row into a spend bucket.

    Priority order: keyword rules on description, then category-based
    fallback, then price heuristic for capital equipment, then default
    to 'Parts / Materials'.
    """
    desc = str(description).strip()
    cat = str(product_category).strip()

    if total_spend < 0:
        return "Discount / Credit"

    for bucket, pattern in _BUCKET_RULES:
        if pattern.search(desc):
            return bucket

    if cat in _CATEGORY_BUCKET_MAP:
        return _CATEGORY_BUCKET_MAP[cat]

    if avg_unit_price >= CAPITAL_EQUIPMENT_THRESHOLD:
        return "Capital Equipment"

    return "Parts / Materials"

PROJECT_TO_LINE_PREFIX: dict[str, list[str]] = {
    "BF1-Module Line 1": ["BASE1-MOD1", "BASE1-CELL1"],
    "BF1-Module Line 2": ["BASE1-MOD2", "BASE1-CELL2"],
    "BF1-Inverter Line 1": ["BASE1-INV1"],
}

STATION_KEYWORD_MAP: dict[str, list[str]] = {
    "ST10000": ["cell pallet", "pallet unload", "cell prep"],
    "ST11000": ["tim dispense", "heatsink tim"],
    "ST12000": ["adhesive dispense"],
    "ST13000": ["pcba press"],
    "ST14000": ["pcba fasten"],
    "ST15000": ["hipot", "heatsink hipot"],
    "ST22000": ["current collector weld", "laser weld", "trufiber", "precitec", "lwm"],
    "ST24000": ["ground bond"],
    "ST25000": ["bms calibration", "pre-fsw functional"],
    "ST31000": ["functional test"],
    "ST33000": ["friction stir", "fsw", "enclosure weld"],
    "ST35000": ["leak test", "leakmaster", "leak master"],
    "ST36000": ["leak re-test"],
    "ST40000": ["packout", "pack out", "tray marriage"],
}

OVERALL_INTEGRATION_KEYWORDS: list[str] = [
    "agv", "conveyor", "conveyance", "fleet management", "maestro",
    "scada", "ignition", "plc", "line integration",
]


# ---------------------------------------------------------------------------
# BF1 Station Loading
# ---------------------------------------------------------------------------

def load_bf1_stations(xlsx_path: str | Path) -> tuple[list[dict], list[dict]]:
    """Load station list and cost breakdown from the BF1 planning Excel.

    Returns (stations, cost_breakdown) where each is a list of dicts.
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)

    stations: list[dict] = []
    ws_overall = wb["BF1 PROD Overall"]
    rows = list(ws_overall.iter_rows(values_only=True))
    for row in rows[1:]:
        sid = str(row[0]).strip() if row[0] else ""
        if not sid or not sid.startswith("BASE"):
            continue
        stations.append({
            "station_id": sid,
            "process_name": str(row[1]).strip() if row[1] else "",
            "station_type": str(row[2]).strip() if row[2] else "",
            "owner": str(row[4]).strip() if row[4] else "",
            "vendor": str(row[6]).strip() if row[6] else "",
            "status": str(row[10]).strip() if row[10] else "",
            "forecasted_cost": float(row[11]) if row[11] is not None else 0.0,
        })

    cost_breakdown: list[dict] = []
    ws_cb = wb["BF1 PROD Cost Breakdown"]
    cb_rows = list(ws_cb.iter_rows(values_only=True))
    for row in cb_rows[1:]:
        sid = str(row[0]).strip() if row[0] else ""
        if not sid or not sid.startswith("BASE"):
            continue
        cost_breakdown.append({
            "station_id": sid,
            "process_name": str(row[1]).strip() if row[1] else "",
            "equipment": str(row[2]).strip() if row[2] else "",
            "owner": str(row[3]).strip() if row[3] else "",
            "unit_cost": float(row[5]) if row[5] is not None else 0.0,
            "total_cost": float(row[6]) if row[6] is not None else 0.0,
            "vendor": str(row[8]).strip() if row[8] else "",
        })

    wb.close()
    return stations, cost_breakdown


# ---------------------------------------------------------------------------
# Product Category Splitting
# ---------------------------------------------------------------------------

def split_product_category(df: pd.DataFrame) -> pd.DataFrame:
    """Parse line_description to extract product_category and item_description."""
    out = df.copy()
    cats: list[str] = []
    items: list[str] = []

    for desc in out["line_description"].fillna(""):
        matched = False
        for cat in KNOWN_CATEGORIES:
            if desc.startswith(cat):
                cats.append(cat)
                items.append(desc[len(cat):].strip())
                matched = True
                break
        if not matched:
            cats.append("")
            items.append(desc)

    out["product_category"] = cats
    out["item_description"] = items
    return out


# ---------------------------------------------------------------------------
# Section Header Merging
# ---------------------------------------------------------------------------

def merge_section_headers(df: pd.DataFrame) -> pd.DataFrame:
    """For rows where item_description is blank after category split,
    pull the description from the previous zero-qty row (same PO, seq-1)."""
    out = df.copy()
    needs_fill = (
        (out["item_description"] == "")
        & (out["product_category"] != "")
        & (out["product_qty"] != 0)
    )
    if not needs_fill.any():
        return out

    for idx in out.index[needs_fill]:
        row = out.loc[idx]
        po = row["po_number"]
        seq = row.get("line_sequence")
        if pd.isna(seq):
            continue
        prev_mask = (
            (out["po_number"] == po)
            & (out["line_sequence"] == seq - 1)
            & (out["product_qty"] == 0)
        )
        prev = out.loc[prev_mask]
        if not prev.empty:
            out.at[idx, "item_description"] = prev.iloc[0]["line_description"]

    return out


# ---------------------------------------------------------------------------
# Line Type Classification
# ---------------------------------------------------------------------------

def classify_line_type(df: pd.DataFrame) -> pd.DataFrame:
    """Tag each row with a line_type: spend, section_header, payment_terms, misc."""
    out = df.copy()
    types: list[str] = []

    for _, row in out.iterrows():
        source = row.get("source", "odoo")
        qty = row.get("product_qty", 0)
        desc = str(row.get("line_description", ""))

        if source == "ramp":
            types.append("spend")
            continue

        if qty != 0:
            types.append("spend")
        elif _PAYMENT_TERMS_PATTERN.search(desc):
            types.append("payment_terms")
        elif qty == 0:
            types.append("section_header")
        else:
            types.append("misc")

    out["line_type"] = types
    return out


# ---------------------------------------------------------------------------
# CAPEX Flag
# ---------------------------------------------------------------------------

def tag_capex_flag(df: pd.DataFrame) -> pd.DataFrame:
    """Add boolean is_capex column."""
    out = df.copy()
    out["is_capex"] = out["product_category"].isin(CAPEX_CATEGORIES)
    return out


# ---------------------------------------------------------------------------
# Ramp Normalization
# ---------------------------------------------------------------------------

def load_and_normalize_ramp(csv_path: str | Path) -> pd.DataFrame:
    """Load Ramp CSV, filter to CAPEX+materials categories, reshape to Odoo schema."""
    import hashlib

    ramp = pd.read_csv(str(csv_path), encoding="utf-8-sig")

    ramp = ramp[ramp["Accounting Category"].isin(RAMP_TO_ODOO_CATEGORY.keys())].copy()
    ramp = ramp.reset_index(drop=True)

    n = len(ramp)
    out = pd.DataFrame(index=range(n))
    out["source"] = "ramp"

    def _stable_id(row: pd.Series, prefix: str) -> str:
        key = f"{row.get('Transaction Date','')}|{row.get('Merchant Name','')}|{row.get('Amount','')}|{row.get('User','')}"
        return prefix + hashlib.md5(key.encode()).hexdigest()[:8]

    out["po_number"] = [_stable_id(ramp.iloc[i], "RAMP-") for i in range(n)]
    out["date_order"] = pd.to_datetime(
        ramp["Transaction Date"].values, format="%m/%d/%y"
    ).strftime("%Y-%m-%d")
    out["po_state"] = "purchase"
    out["po_invoice_status"] = ""
    out["po_receipt_status"] = ""
    out["vendor_name"] = ramp["Merchant Name"].values
    out["vendor_ref"] = ""
    out["line_description"] = ramp["Merchant Name"].values
    out["product_category"] = ramp["Accounting Category"].map(RAMP_TO_ODOO_CATEGORY).values
    out["item_description"] = (ramp["Merchant Name"] + " (CC)").values

    out["product_id"] = ""
    out["product_qty"] = ""
    out["qty_received"] = ""
    out["product_uom"] = ""
    out["price_unit"] = ""
    out["price_subtotal"] = ramp["Amount"].values
    out["price_tax"] = 0.0
    out["price_total"] = ramp["Amount"].values
    out["line_date_planned"] = ""
    out["line_sequence"] = ""
    out["line_id"] = [_stable_id(ramp.iloc[i], "RL-") for i in range(n)]

    proj = ramp["Accounting Projects"].fillna("")
    out["project_name"] = proj.str.replace(" - Base Power, Inc.", "", regex=False).str.strip().values

    user = ramp["User"].copy()
    for alias_from, alias_to in RAMP_USER_ALIASES.items():
        user = user.replace(alias_from, alias_to)
    out["created_by_name"] = user.values

    out["po_amount_total"] = ramp["Amount"].values
    out["bill_count"] = 0
    out["bill_amount_total"] = 0.0
    out["bill_amount_paid"] = 0.0
    out["bill_amount_open"] = 0.0
    out["bill_payment_status"] = ""
    out["po_notes"] = ""
    out["is_capex"] = out["product_category"].isin(CAPEX_CATEGORIES)

    out["ramp_card"] = ramp["Card Display Name"].values
    out["ramp_department"] = ramp["Ramp Department"].values
    out["ramp_location"] = ramp["Ramp Location"].fillna("").values
    out["ramp_category"] = ramp["Ramp Category"].fillna("").values
    out["ramp_merchant"] = ramp["Accounting Merchant"].fillna("").values

    out["item_description"] = ramp.apply(
        lambda r: "{merchant} | {user} | {date} | {card} | {dept}".format(
            merchant=r.get("Merchant Name", ""),
            user=r.get("User", ""),
            date=r.get("Transaction Date", ""),
            card=r.get("Card Display Name", ""),
            dept=r.get("Ramp Department", ""),
        ), axis=1,
    ).values
    out["line_description"] = out["item_description"].values

    return out


# ---------------------------------------------------------------------------
# Part Number Extraction
# ---------------------------------------------------------------------------

_PART_NUMBER_PATTERNS: list[tuple[str, re.Pattern]] = [
    ("prd-BPC", re.compile(r"(prd-BPC-\d{5,})", re.IGNORECASE)),
    ("PX", re.compile(r"(PX\s*\d{5,})")),
    ("Bosch Rexroth", re.compile(r"\b(384\d{7})\b")),
    ("BCC", re.compile(r"\b(BCC\w{4,})\b")),
    ("BN", re.compile(r"\bBN\s+(\w{5,})\b")),
    ("SC", re.compile(r"\bSC\s+(\w{8,})\b")),
    ("Siemens", re.compile(r"\b(6\w{2}\d{4,}\w*)\b")),
    ("Model", re.compile(
        r"\b(TruFiber\s*\d+[A-Z]*|R-\d{3,}iA[/\w]*|LR-?\d+iA[/\w]*"
        r"|M-\d{3,}iC[/\w]*|GDP\d+\w*|VHX-\w+|GL-\w{3,}|DT\d{4}"
        r"|MGB2?-\w+|IS\d{4}\w*|RM\d{4}|BT\d{4}\w*)\b"
    )),
]


def extract_part_numbers(description: str) -> str:
    """Extract structured part/model identifiers from a description string.
    Returns a JSON string of [{type, value}, ...].
    """
    if not description or pd.isna(description):
        return "[]"
    results: list[dict[str, str]] = []
    seen: set[str] = set()
    for ptype, pattern in _PART_NUMBER_PATTERNS:
        for m in pattern.finditer(str(description)):
            val = m.group(1) if m.lastindex else m.group(0)
            val = val.strip()
            if val not in seen:
                seen.add(val)
                results.append({"type": ptype, "value": val})
    return json.dumps(results) if results else "[]"


# ---------------------------------------------------------------------------
# 3-Tier Station Auto-Mapping Agent
# ---------------------------------------------------------------------------

def _build_vendor_station_lookup(
    cost_breakdown: list[dict],
) -> dict[str, list[str]]:
    """Build a lowercase-vendor -> list-of-station_ids lookup from cost breakdown."""
    lookup: dict[str, list[str]] = {}
    for row in cost_breakdown:
        vendor = row["vendor"].lower().strip()
        sid = row["station_id"]
        if not vendor or not sid:
            continue
        lookup.setdefault(vendor, [])
        if sid not in lookup[vendor]:
            lookup[vendor].append(sid)
    return lookup


def _fuzzy_vendor_match(
    vendor_name: str,
    vendor_lookup: dict[str, list[str]],
) -> list[str]:
    """Find station IDs whose cost-breakdown vendor fuzzy-matches the PO vendor."""
    vn = vendor_name.lower().strip()
    matched_stations: list[str] = []
    for cb_vendor, sids in vendor_lookup.items():
        if cb_vendor in vn or vn in cb_vendor:
            matched_stations.extend(sids)
        elif len(cb_vendor) > 3 and len(vn) > 3:
            cb_words = set(cb_vendor.split())
            vn_words = set(vn.split())
            if cb_words & vn_words and len(cb_words & vn_words) >= 1:
                first_cb = cb_vendor.split()[0]
                first_vn = vn.split()[0]
                if first_cb == first_vn or first_cb in vn or first_vn in cb_vendor:
                    matched_stations.extend(sids)
    return list(set(matched_stations))


def _keyword_station_match(description: str) -> list[str]:
    """Match description keywords to station ID suffixes."""
    desc_lower = description.lower()
    matched: list[str] = []

    for kw in OVERALL_INTEGRATION_KEYWORDS:
        if kw in desc_lower:
            matched.append("OVERALL")
            break

    for suffix, keywords in STATION_KEYWORD_MAP.items():
        for kw in keywords:
            if kw in desc_lower:
                matched.append(suffix)
                break

    return list(set(matched))


def _pick_best_station(
    candidates: dict[str, int],
    line_prefixes: list[str],
    all_stations: list[dict],
) -> tuple[str, str, int]:
    """Pick the best station from scored candidates, preferring ones matching line prefixes.
    Returns (station_id, station_name, score).
    """
    if not candidates:
        return "", "", 0

    station_name_map = {s["station_id"]: s["process_name"] for s in all_stations}

    best_id = ""
    best_score = 0
    for sid, score in candidates.items():
        if score > best_score:
            best_score = score
            best_id = sid
        elif score == best_score and best_id:
            for pfx in line_prefixes:
                if sid.startswith(pfx) and not best_id.startswith(pfx):
                    best_id = sid
                    break

    return best_id, station_name_map.get(best_id, ""), best_score


def _project_allows_base2(project_name: str) -> bool:
    """Only allow BASE2 station assignment when project explicitly indicates BASE2/BF2."""
    p = str(project_name).upper()
    return ("BASE2" in p) or ("BF2" in p) or ("CIP-BF2-" in p)


def auto_map_stations(
    df: pd.DataFrame,
    stations: list[dict],
    cost_breakdown: list[dict],
) -> pd.DataFrame:
    """Run the 3-tier station mapping agent on the unified dataframe."""
    out = df.copy()
    vendor_lookup = _build_vendor_station_lookup(cost_breakdown)
    station_name_map = {s["station_id"]: s["process_name"] for s in stations}
    station_ids_set = set(s["station_id"] for s in stations)

    result_station_id: list[str] = []
    result_station_name: list[str] = []
    result_confidence: list[str] = []
    result_reason: list[str] = []

    for _, row in out.iterrows():
        proj = str(row.get("project_name", "")).strip()
        vendor = str(row.get("vendor_name", "")).strip()
        creator = str(row.get("created_by_name", "")).strip()
        desc = str(row.get("item_description", "")).strip()
        line_desc = str(row.get("line_description", "")).strip()
        source = row.get("source", "odoo")
        line_type = row.get("line_type", "spend")
        ramp_card = str(row.get("ramp_card", "")).strip()
        date_order = pd.to_datetime(row.get("date_order", ""), errors="coerce")
        if proj.lower() == "nan":
            proj = ""
        if vendor.lower() == "nan":
            vendor = ""
        if creator.lower() == "nan":
            creator = ""
        if desc.lower() == "nan":
            desc = ""
        if line_desc.lower() == "nan":
            line_desc = ""
        if ramp_card.lower() == "nan":
            ramp_card = ""

        if line_type != "spend":
            result_station_id.append("")
            result_station_name.append("")
            result_confidence.append("none")
            result_reason.append("non-spend row")
            continue

        # --- Tier 3: Non-PROD / Pilot auto-exclusion ---
        if proj in NON_PROD_PROJECTS:
            tag = "pilot_npi" if "NPI" in proj or "Pilot" in proj or "Prototype" in proj else "non_prod"
            result_station_id.append("")
            result_station_name.append("")
            result_confidence.append("high")
            result_reason.append(f"{tag}: project={proj}")
            continue

        if source == "ramp" and ramp_card in PILOT_RAMP_CARDS:
            result_station_id.append("")
            result_station_name.append("")
            result_confidence.append("medium")
            result_reason.append(f"pilot_npi: ramp card='{ramp_card}'")
            continue

        # --- Tier 3b: Ramp routing when project context is missing ---
        # This prevents persistent "unmapped" buckets for operational card spend.
        if source == "ramp" and not proj:
            if creator in {"Christopher George", "Kelsea Allenbaugh"}:
                result_station_id.append("BF1-Facilities and Infrastructure")
                result_station_name.append("BF1-Facilities and Infrastructure")
                result_confidence.append("high")
                result_reason.append(f"non_prod: ramp creator={creator}")
                continue
            if pd.notna(date_order) and date_order < pd.Timestamp("2025-11-01"):
                result_station_id.append("BF1-NPI & Pilot Equipment")
                result_station_name.append("BF1-NPI & Pilot Equipment")
                result_confidence.append("medium")
                result_reason.append("pilot_npi: ramp pre-2025-11-01")
                continue
            result_station_id.append("BF1-Other Allocation")
            result_station_name.append("BF1-Other Allocation")
            result_confidence.append("low")
            result_reason.append("non_prod: ramp no project -> other allocation")
            continue

        # --- Tier 1: Direct CIP project mapping ---
        if proj.startswith("CIP-BF1-"):
            code = proj.split(" : ")[0]
            sid = code.replace("CIP-BF1-", "BASE1-")
            if sid in station_ids_set:
                result_station_id.append(sid)
                result_station_name.append(station_name_map.get(sid, ""))
                result_confidence.append("high")
                result_reason.append(f"CIP project direct map: {proj}")
                continue
            for real_sid in station_ids_set:
                if real_sid.startswith(sid):
                    result_station_id.append(real_sid)
                    result_station_name.append(station_name_map.get(real_sid, ""))
                    result_confidence.append("high")
                    result_reason.append(f"CIP project prefix map: {proj}")
                    break
            else:
                result_station_id.append(sid)
                result_station_name.append(station_name_map.get(sid, code))
                result_confidence.append("medium")
                result_reason.append(f"CIP project (station not in master): {proj}")
                continue
            continue

        # --- Tier 2: Scored matching (vendor + project-line + keywords) ---
        line_prefixes: list[str] = PROJECT_TO_LINE_PREFIX.get(proj, [])
        candidates: dict[str, int] = {}

        vendor_stations = _fuzzy_vendor_match(vendor, vendor_lookup)
        # Guardrail: if project maps to a production line, vendor matches must stay
        # within that line family; otherwise high vendor score can mis-route to BASE2.
        if line_prefixes:
            vendor_stations = [
                sid for sid in vendor_stations
                if any(sid.startswith(pfx) for pfx in line_prefixes)
            ]
        # Guardrail: BASE2 assignments require explicit project intent.
        if not _project_allows_base2(proj):
            vendor_stations = [sid for sid in vendor_stations if not sid.startswith("BASE2-")]
        for sid in vendor_stations:
            candidates[sid] = candidates.get(sid, 0) + 3

        if line_prefixes:
            for sid in station_ids_set:
                for pfx in line_prefixes:
                    if sid.startswith(pfx):
                        candidates[sid] = candidates.get(sid, 0) + 2
                        break
        if not _project_allows_base2(proj):
            candidates = {sid: sc for sid, sc in candidates.items() if not sid.startswith("BASE2-")}

        search_text = f"{desc} {line_desc}".lower()
        kw_matches = _keyword_station_match(search_text)
        for kw_suffix in kw_matches:
            if kw_suffix == "OVERALL":
                for pfx in (line_prefixes or ["BASE1-MOD1", "BASE1-MOD2", "BASE1-INV1"]):
                    base = pfx
                    if base in station_ids_set:
                        candidates[base] = candidates.get(base, 0) + 1
            else:
                for sid in station_ids_set:
                    if kw_suffix in sid:
                        if line_prefixes:
                            for pfx in line_prefixes:
                                if sid.startswith(pfx):
                                    candidates[sid] = candidates.get(sid, 0) + 1
                        else:
                            candidates[sid] = candidates.get(sid, 0) + 1

        if line_prefixes and not candidates:
            for sid in station_ids_set:
                for pfx in line_prefixes:
                    if sid.startswith(pfx):
                        candidates[sid] = candidates.get(sid, 0) + 1

        best_sid, best_name, best_score = _pick_best_station(
            candidates, line_prefixes, stations
        )

        if best_score >= 5:
            conf = "high"
        elif best_score >= 3:
            conf = "medium"
        elif best_score >= 1:
            conf = "low"
        else:
            conf = "none"

        reasons: list[str] = []
        if vendor_stations:
            reasons.append(f"vendor='{vendor}'")
        if line_prefixes:
            reasons.append(f"project->line={line_prefixes}")
        if kw_matches:
            reasons.append(f"keywords={kw_matches}")

        result_station_id.append(best_sid)
        result_station_name.append(best_name)
        result_confidence.append(conf)
        result_reason.append(
            f"score={best_score}: {'; '.join(reasons)}" if reasons else "no signals"
        )

    out["station_id"] = result_station_id
    out["station_name"] = result_station_name
    out["mapping_confidence"] = result_confidence
    out["mapping_reason"] = result_reason

    return out


# ---------------------------------------------------------------------------
# Apply Human Overrides
# ---------------------------------------------------------------------------

def apply_overrides(
    df: pd.DataFrame,
    overrides_source: str | Path | dict,
    stations: list[dict],
) -> pd.DataFrame:
    """Apply station_overrides.json on top of agent mappings.

    ``overrides_source`` can be a file path (str/Path) or a pre-loaded dict.
    """
    if isinstance(overrides_source, dict):
        overrides = overrides_source
    else:
        path = Path(overrides_source)
        if not path.exists():
            df["mapping_status"] = df["mapping_confidence"].apply(
                lambda c: "auto" if c in ("high", "medium", "low") else "unmapped"
            )
            return df
        with open(path, encoding="utf-8") as f:
            overrides = json.load(f)

    if not overrides:
        df["mapping_status"] = df["mapping_confidence"].apply(
            lambda c: "auto" if c in ("high", "medium", "low") else "unmapped"
        )
        return df

    station_name_map = {s["station_id"]: s["process_name"] for s in stations}
    project_code_set = {
        "BF1-NPI & Pilot Equipment", "BF1-Prototype R&D Lines",
        "BF1-Quality Equipment", "BF1-Facilities and Infrastructure",
        "BF1-Manufacturing IT Systems", "BF1-Warehousing and Material Handling",
        "BF1-Maintenance and Spares", "BF1-Other Allocation", "BF1-Module Line 1",
        "BF1-Module Line 2", "BF1-Inverter Line 1",
    }
    out = df.copy()

    statuses: list[str] = []
    for _, row in out.iterrows():
        lid = str(row.get("line_id", ""))
        if lid in overrides:
            ov = overrides[lid]
            status = ov.get("status", "confirmed")
            sid = ov.get("station_id", "")
            if sid:
                out.at[row.name, "station_id"] = sid
                if sid in station_name_map:
                    out.at[row.name, "station_name"] = station_name_map[sid]
                elif sid in project_code_set:
                    out.at[row.name, "station_name"] = sid
                    out.at[row.name, "mapping_reason"] = f"human override: project={sid}"
            if status == "skip":
                statuses.append("skipped")
            elif status == "non_prod":
                statuses.append("non_prod")
            elif status == "pilot_npi":
                statuses.append("pilot_npi")
            else:
                statuses.append("confirmed" if sid else "overridden")
        else:
            conf = row.get("mapping_confidence", "none")
            if conf in ("high", "medium", "low"):
                statuses.append("auto")
            else:
                statuses.append("unmapped")

    out["mapping_status"] = statuses
    return out
